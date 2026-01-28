"""
build_priser_fixing_selector.py

Creates an Excel workbook with:
- Dropdown selection for FixingTime and Date
- Formulas that populate:
    EURNOK Days -> T43:T46
    USDNOK Days -> AA43:AA46

Input ("tidy") must contain columns:
  FixingTime, Pair, Date, Days, Points

Output:
  O:\\MM Internbanken\\Samba\\Priser_fixing_selector.xlsx

Requirements:
  pip install openpyxl pandas
"""

from __future__ import annotations

from pathlib import Path
import sys
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


# =========================
# CONFIG - EDIT THIS
# =========================
SRC_PATH = Path(r"O:\MM Internbanken\Samba\Priser_tidy.xlsx")  # or .csv
OUT_PATH = Path(r"O:\MM Internbanken\Samba\Priser_fixing_selector.xlsx")


# =========================
# HELPERS
# =========================
REQUIRED_COLS = ["FixingTime", "Pair", "Date", "Days", "Points"]

ALLOWED_FIXINGS = ["08:00", "08:30", "09:00", "09:30", "10:00", "10:30"]
ALLOWED_SET = set(ALLOWED_FIXINGS)


def _fix_sort_key(t: str):
    try:
        hh, mm = t.split(":")
        return int(hh), int(mm)
    except Exception:
        return (999, 999)


def _read_tidy_source(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Source file not found: {path}")

    if path.suffix.lower() in [".xlsx", ".xlsm", ".xls"]:
        wb = load_workbook(path, data_only=True)
        # Prefer a sheet named "Tidy" if it exists, otherwise first sheet
        sheet_name = "Tidy" if "Tidy" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.values)
        if not rows or len(rows) < 2:
            raise RuntimeError(f"Sheet '{sheet_name}' appears empty in: {path}")

        headers = [str(x).strip() if x is not None else "" for x in rows[0]]
        df = pd.DataFrame(rows[1:], columns=headers)

    elif path.suffix.lower() in [".csv", ".txt"]:
        # Support comma decimal in Points if present
        df = pd.read_csv(path, sep=None, engine="python", dtype=str)
    else:
        raise RuntimeError(f"Unsupported input format: {path.suffix} (use .xlsx or .csv)")

    # Normalize column names (strip)
    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in REQUIRED_COLS if c not in df.columns]
    if missing:
        raise RuntimeError(
            f"Missing required columns: {missing}\n"
            f"Found columns: {list(df.columns)}"
        )

    # Normalize types
    df["FixingTime"] = df["FixingTime"].astype(str).str.strip()
    df["Pair"] = df["Pair"].astype(str).str.strip().str.upper()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date
    df["Days"] = pd.to_numeric(df["Days"], errors="coerce")

    # Points may have comma decimal; normalize robustly (not strictly needed for Days output)
    pts_raw = df["Points"].astype(str).str.strip()
    pts_raw = pts_raw.str.replace(" ", "", regex=False)
    pts_raw = pts_raw.str.replace(",", ".", regex=False)
    df["Points"] = pd.to_numeric(pts_raw, errors="coerce")

    df = df.dropna(subset=["FixingTime", "Pair", "Date", "Days"])

    # Keep only the fixings we care about (optional but recommended)
    df = df[df["FixingTime"].isin(ALLOWED_SET)]

    if df.empty:
        raise RuntimeError(
            "After cleaning, no rows remain. Check that FixingTime values match "
            f"{ALLOWED_FIXINGS} and that Date/Days are valid."
        )

    df["DateText"] = df["Date"].astype(str)  # YYYY-MM-DD
    return df


def _build_selector_workbook(df: pd.DataFrame, out_path: Path) -> None:
    """
    Build output workbook using:
      - LISTS sheet for dropdown sources
      - DATA sheet with lookup keys
      - OUTPUT sheet with dropdowns + VLOOKUP formulas
    """
    df = df.copy()

    # Rank the four smallest day buckets per FixingTime/Date/Pair
    df = df.sort_values(["FixingTime", "DateText", "Pair", "Days"])
    df["Rank"] = df.groupby(["FixingTime", "DateText", "Pair"]).cumcount() + 1
    df = df[df["Rank"].between(1, 4)].copy()

    # Key for VLOOKUP
    df["Key"] = (
        df["FixingTime"]
        + "|"
        + df["DateText"]
        + "|"
        + df["Pair"]
        + "|"
        + df["Rank"].astype(int).astype(str)
    )

    fixings = sorted(df["FixingTime"].unique().tolist(), key=_fix_sort_key)
    dates = sorted(df["DateText"].unique().tolist())

    if not fixings:
        raise RuntimeError("No fixing times found after preprocessing.")
    if not dates:
        raise RuntimeError("No dates found after preprocessing.")

    wb = Workbook()
    wb.remove(wb.active)

    ws_out = wb.create_sheet("OUTPUT")
    ws_data = wb.create_sheet("DATA")
    ws_lists = wb.create_sheet("LISTS")

    bold = Font(bold=True)

    # -----------------
    # LISTS
    # -----------------
    ws_lists["A1"] = "FixingTime"
    ws_lists["A1"].font = bold
    for i, ft in enumerate(fixings, start=2):
        ws_lists[f"A{i}"] = ft

    ws_lists["C1"] = "DateText"
    ws_lists["C1"].font = bold
    for i, dt in enumerate(dates, start=2):
        ws_lists[f"C{i}"] = dt

    ws_lists.column_dimensions["A"].width = 14
    ws_lists.column_dimensions["C"].width = 14

    # -----------------
    # OUTPUT (top-left)
    # -----------------
    ws_out["A1"] = "Select Fixing time (B2) and Date (B3). Days populate automatically."
    ws_out["A1"].font = bold

    ws_out["A2"] = "Fixing time"
    ws_out["A2"].font = bold
    ws_out["B2"] = fixings[-1]

    ws_out["A3"] = "Date"
    ws_out["A3"].font = bold
    ws_out["B3"] = dates[-1]

    # Visible area (optional)
    ws_out["D2"] = "EURNOK Days (visible)"
    ws_out["D2"].font = bold
    ws_out["J2"] = "USDNOK Days (visible)"
    ws_out["J2"].font = bold

    for r in range(4):
        rank = r + 1
        ws_out[f"D{4+r}"] = (
            f'=IFERROR(VLOOKUP($B$2&"|"&$B$3&"|EURNOK|{rank}",DATA!$A:$F,6,FALSE),"")'
        )
        ws_out[f"J{4+r}"] = (
            f'=IFERROR(VLOOKUP($B$2&"|"&$B$3&"|USDNOK|{rank}",DATA!$A:$F,6,FALSE),"")'
        )

    # Your requested mapping
    ws_out["T42"] = "EURNOK Days"
    ws_out["T42"].font = bold
    ws_out["AA42"] = "USDNOK Days"
    ws_out["AA42"].font = bold

    for r in range(4):
        rank = r + 1
        ws_out[f"T{43+r}"] = (
            f'=IFERROR(VLOOKUP($B$2&"|"&$B$3&"|EURNOK|{rank}",DATA!$A:$F,6,FALSE),"")'
        )
        ws_out[f"AA{43+r}"] = (
            f'=IFERROR(VLOOKUP($B$2&"|"&$B$3&"|USDNOK|{rank}",DATA!$A:$F,6,FALSE),"")'
        )

    for col, w in {"A":16, "B":14, "D":22, "J":22, "T":16, "AA":16}.items():
        ws_out.column_dimensions[col].width = w
    ws_out.freeze_panes = "A4"

    # Dropdowns
    dv_fix = DataValidation(
        type="list",
        formula1=f"=LISTS!$A$2:$A${1 + len(fixings)}",
        allow_blank=False,
        showDropDown=True,
    )
    dv_date = DataValidation(
        type="list",
        formula1=f"=LISTS!$C$2:$C${1 + len(dates)}",
        allow_blank=False,
        showDropDown=True,
    )
    ws_out.add_data_validation(dv_fix)
    ws_out.add_data_validation(dv_date)
    dv_fix.add(ws_out["B2"])
    dv_date.add(ws_out["B3"])

    # -----------------
    # DATA
    # -----------------
    ws_data.append(["Key", "FixingTime", "DateText", "Pair", "Rank", "Days"])
    for cell in ws_data[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    for _, row in df[["Key", "FixingTime", "DateText", "Pair", "Rank", "Days"]].iterrows():
        ws_data.append(
            [
                row["Key"],
                row["FixingTime"],
                row["DateText"],
                row["Pair"],
                int(row["Rank"]),
                float(row["Days"]),
            ]
        )

    widths = [42, 12, 12, 10, 8, 10]
    for i, w in enumerate(widths, start=1):
        ws_data.column_dimensions[get_column_letter(i)].width = w
    ws_data.freeze_panes = "A2"

    # Sanity check: ensure formulas exist
    for addr in ["D4", "J4", "T43", "AA43"]:
        v = ws_out[addr].value
        if not (isinstance(v, str) and v.startswith("=")):
            raise RuntimeError(f"Expected a formula in {addr}, found: {v!r}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    print(f"[INFO] Reading source: {SRC_PATH}")
    df = _read_tidy_source(SRC_PATH)

    print(f"[INFO] Rows after cleaning: {len(df)}")
    print(f"[INFO] Writing output: {OUT_PATH}")
    _build_selector_workbook(df, OUT_PATH)

    print("[OK] Done.")
    print("Open the output file and use OUTPUT!B2 (Fixing time) and OUTPUT!B3 (Date).")
    print("EURNOK Days populate in OUTPUT!T43:T46 and USDNOK Days in OUTPUT!AA43:AA46.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        raise
