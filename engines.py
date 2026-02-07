"""
Data engines for Onyx Terminal.
Contains ExcelEngine and BloombergEngine.
"""
import threading
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

from config import (
    BASE_HISTORY_PATH, DAY_FILES, RECON_FILE, WEIGHTS_FILE,
    RECON_MAPPING, DAYS_MAPPING, RULES_DB, SWET_CM_RECON_MAPPING,
    WEIGHTS_FILE_CELLS, WEIGHTS_MODEL_CELLS, USE_MOCK_DATA,
    EXCEL_CM_RATES_MAPPING, RECON_CELL_MAPPING, get_logger,
    INTERNAL_BASKET_MAPPING, WEIGHTS_COLS, NIBOR_FIXING_TICKERS,
    DATA_DIR,
)
from nibor_file_manager import get_nibor_file_path, NiborFileManager
from file_watcher import FileWatcher, compute_file_hash

log = get_logger("engines")
from utils import copy_to_cache_fast, safe_float, to_date

# Bloomberg API optional
try:
    import blpapi
except ImportError:
    blpapi = None

# xlwings optional (Windows only — reads from running Excel instance)
try:
    import xlwings as xw
    _HAS_XLWINGS = True
except ImportError:
    xw = None
    _HAS_XLWINGS = False


def build_required_cell_set() -> set[tuple[int, int]]:
    """Build set of all required cells for Excel reading."""
    needed = set()

    for cell, _, _ in RECON_MAPPING:
        needed.add(coordinate_to_tuple(cell))

    for cell, _, _ in DAYS_MAPPING:
        needed.add(coordinate_to_tuple(cell))

    for _, top_cell, ref_cell, logic, _ in RULES_DB:
        if top_cell and top_cell != "-":
            needed.add(coordinate_to_tuple(top_cell))
        if ref_cell and ref_cell != "-" and logic in ("Exakt Match", "Avrundat 2 dec", "Minimum"):
            needed.add(coordinate_to_tuple(ref_cell))

    for cell, _, _ in SWET_CM_RECON_MAPPING:
        needed.add(coordinate_to_tuple(cell))

    for c in WEIGHTS_MODEL_CELLS.values():
        needed.add(coordinate_to_tuple(c))

    # Add reconciliation cells for drawer (K30-K33, M30-M33, etc.)
    for field_cells in RECON_CELL_MAPPING.values():
        for cell in field_cells.values():
            needed.add(coordinate_to_tuple(cell))

    return needed


REQUIRED_CELLS = build_required_cell_set()


def _open_workbook(file_path: Path, retries: int = 3, delay: float = 0.5, **kwargs):
    """Open an Excel workbook with retry logic for locked files.

    Retries reading the file directly before falling back to a cache copy.

    Args:
        file_path: Path to the Excel file
        retries: Number of retry attempts when file is locked
        delay: Seconds to wait between retries
        **kwargs: Additional arguments for load_workbook (e.g. data_only, read_only)

    Returns:
        Tuple of (workbook, used_cache: bool)
    """
    defaults = {"data_only": True, "read_only": True}
    defaults.update(kwargs)

    # Try direct read with retries
    for attempt in range(retries):
        try:
            return load_workbook(file_path, **defaults), False
        except (PermissionError, OSError):
            if attempt < retries - 1:
                log.debug(f"[_open_workbook] File locked, retry {attempt + 1}/{retries} in {delay}s...")
                time.sleep(delay)

    # All retries exhausted — fall back to cache copy
    log.warning(f"[_open_workbook] File locked after {retries} retries, using cache copy")
    temp_path = copy_to_cache_fast(file_path)
    return load_workbook(temp_path, **defaults), True


def _read_cells_xlwings(file_path: Path, required_cells: set[tuple[int, int]],
                        cm_mapping: dict[str, str]) -> tuple[dict, dict, str] | None:
    """Try to read cells directly from a running Excel instance via xlwings.

    Returns (recon_data, cm_rates, sheet_name) on success, or None if
    xlwings is unavailable or the workbook isn't open in Excel.
    """
    if not _HAS_XLWINGS:
        return None

    try:
        filename = file_path.name
        wb = None

        for app in xw.apps:
            for book in app.books:
                if book.name == filename:
                    wb = book
                    break
            if wb:
                break

        if wb is None:
            log.debug(f"[xlwings] Workbook {filename} not open in Excel")
            return None

        # Choose the correct sheet (prefer latest date-formatted sheet)
        import re
        date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
        date_sheets = [s.name for s in wb.sheets if date_pattern.match(s.name)]
        if date_sheets:
            sheet_name = sorted(date_sheets)[-1]
        else:
            sheet_name = wb.sheets[-1].name

        ws = wb.sheets[sheet_name]

        # Read required cells (row, col are 1-indexed — xlwings uses the same)
        recon = {}
        for (r, c) in required_cells:
            recon[(r, c)] = ws.range((r, c)).value

        # Read CM rates
        cm_rates = {}
        for key, cell_ref in cm_mapping.items():
            val = ws.range(cell_ref).value
            cm_rates[key] = safe_float(val, None)

        log.info(f"[xlwings] Read {len(recon)} cells + {len(cm_rates)} CM rates from live Excel ({sheet_name})")
        return recon, cm_rates, sheet_name

    except Exception as e:
        log.debug(f"[xlwings] Failed to read from Excel: {e}")
        return None


def _parse_date_cell(cell_value) -> datetime | None:
    """Parse a cell value as a datetime. Returns None on failure."""
    if isinstance(cell_value, datetime):
        return cell_value
    try:
        return datetime.strptime(str(cell_value), "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _parse_weights_row(usd_val, eur_val, nok_val) -> dict[str, float]:
    """Parse weight values from a row, with sensible defaults."""
    usd = float(usd_val) if usd_val is not None else 0.0
    eur = float(eur_val) if eur_val is not None else 0.0
    nok = float(nok_val) if nok_val is not None else (1.0 - usd - eur)
    return {"USD": usd, "EUR": eur, "NOK": nok}


class ExcelEngine:
    """Engine for reading and processing Excel files."""

    def __init__(self):
        self.day_data = pd.DataFrame()
        self._day_data_ready = False
        self._day_data_err = None

        self.recon_data = {}
        self.current_filename = ""
        self.current_folder_path = BASE_HISTORY_PATH
        self.current_year_loaded = ""
        self.last_loaded_ts: datetime | None = None

        self._last_src: Path | None = None
        self._last_mtime: float | None = None
        self._last_size: int | None = None
        self._last_hash: str | None = None
        self.used_cache_fallback: bool = False

        # File watcher for automatic change detection
        # Uses quiet period to handle Excel auto-save (waits for file to be stable)
        self._file_watcher = FileWatcher(
            on_change_callback=self._on_file_changed,
            quiet_period_seconds=10.0,       # Wait 10s after last change before notifying
            notification_cooldown_seconds=60.0,  # Max 1 notification per minute
            poll_interval_seconds=5.0,
        )
        self._file_watcher.start()
        self._change_callbacks: list[callable] = []

        # WEIGHTS file cache
        self.weights_ok: bool = False
        self.weights_err: str | None = "Not loaded"
        self.weights_last_loaded_ts: datetime | None = None
        self.weights_cells_raw: dict[str, object] = {}
        self.weights_cells_parsed: dict[str, object] = {}

        # Excel CM rates cache (from Nibor fixing workbook)
        self.excel_cm_rates: dict[str, float] = {}

        threading.Thread(target=self._load_day_files_bg, daemon=True).start()

    def _load_day_files_bg(self):
        """Load NIBOR days files with improved error reporting."""
        try:
            dfs = []
            errors = []
            
            log.info(f"[ExcelEngine] Starting to load {len(DAY_FILES)} day files...")
            
            for f_path in DAY_FILES:
                log.info(f"[ExcelEngine] Checking: {f_path}")
                
                if not f_path.exists():
                    err_msg = f"File not found: {f_path}"
                    log.info(f"[ExcelEngine] [ERROR] {err_msg}")
                    errors.append(err_msg)
                    continue
                
                log.info(f"[ExcelEngine] [OK] File exists, attempting to read...")
                
                try:
                    df = pd.read_excel(f_path, engine="openpyxl")
                    log.info(f"[ExcelEngine] [OK] Successfully read {len(df)} rows from {f_path.name}")
                    log.info(f"[ExcelEngine]   Columns: {list(df.columns)}")
                    dfs.append(df)
                except Exception as e1:
                    log.info(f"[ExcelEngine] [WARN] Direct read failed: {e1}")
                    log.info(f"[ExcelEngine]   Trying cache copy method...")
                    try:
                        temp = copy_to_cache_fast(f_path)
                        df = pd.read_excel(temp, engine="openpyxl")
                        log.info(f"[ExcelEngine] [OK] Successfully read from cache copy")
                        dfs.append(df)
                    except Exception as e2:
                        err_msg = f"Failed to read {f_path.name}: {e2}"
                        log.info(f"[ExcelEngine] [ERROR] {err_msg}")
                        errors.append(err_msg)
                        continue

            if dfs:
                log.info(f"[ExcelEngine] Concatenating {len(dfs)} dataframes...")
                day_data = pd.concat(dfs, ignore_index=True)
                log.info(f"[ExcelEngine] Total rows: {len(day_data)}")
                
                if "date" in day_data.columns:
                    log.info(f"[ExcelEngine] Processing 'date' column...")
                    day_data["date"] = pd.to_datetime(day_data["date"], errors="coerce")
                    day_data["date"] = day_data["date"].dt.normalize()
                    day_data = day_data.dropna(subset=["date"]).sort_values("date")
                    log.info(f"[ExcelEngine] After date processing: {len(day_data)} rows")
                else:
                    log.info(f"[ExcelEngine] [WARN] WARNING: 'date' column not found in data!")
                
                self.day_data = day_data
                log.info(f"[ExcelEngine] [OK] Day data loaded successfully")
                log.info(f"[ExcelEngine] Final columns: {list(self.day_data.columns)}")
            else:
                err_msg = "No day files could be loaded"
                log.info(f"[ExcelEngine] [ERROR] {err_msg}")
                if errors:
                    err_msg += ": " + "; ".join(errors)
                self._day_data_err = err_msg
            
            self._day_data_ready = True
            log.info(f"[ExcelEngine] Day files loading complete. Ready={self._day_data_ready}, Error={self._day_data_err}")
            
        except Exception as e:
            err_msg = f"Fatal error loading day files: {e}"
            log.info(f"[ExcelEngine] [ERROR] {err_msg}")
            self._day_data_err = err_msg
            self._day_data_ready = True

    def resolve_latest_path(self) -> tuple[Path | None, str]:
        """
        Resolve the latest NIBOR fixing workbook path.

        Uses dynamic lookup based on current date and development_mode setting:
        - development_mode=True: Loads "_TEST" suffixed files
        - development_mode=False: Loads production files

        Falls back to static RECON_FILE if dynamic lookup fails.
        """
        try:
            # Use dynamic file lookup - get_nibor_file_path reads mode from settings
            file_path = get_nibor_file_path()  # No mode arg = reads from settings

            if file_path.exists():
                self.current_folder_path = file_path.parent
                self.current_year_loaded = file_path.parent.name
                self.current_filename = file_path.name

                # Get mode for logging
                from settings import get_setting
                dev_mode = get_setting("development_mode", False)
                mode_display = "TEST" if dev_mode else "PROD"
                log.info(f"[ExcelEngine] Resolved NIBOR file ({mode_display}): {file_path.name}")
                return file_path, "OK"

        except FileNotFoundError as e:
            log.warning(f"[ExcelEngine] Dynamic file lookup failed: {e}")
        except Exception as e:
            log.warning(f"[ExcelEngine] Error in dynamic file lookup: {e}")

        # Fallback to static RECON_FILE
        if RECON_FILE.exists():
            self.current_folder_path = RECON_FILE.parent
            self.current_year_loaded = RECON_FILE.parent.name
            self.current_filename = RECON_FILE.name
            log.info(f"[ExcelEngine] Using fallback RECON_FILE: {RECON_FILE.name}")
            return RECON_FILE, "OK"

        return None, "File Not Found"

    def load_weights_file(self) -> bool:
        """Load EXACT cells from Weights.xlsx."""
        try:
            if not WEIGHTS_FILE.exists():
                self.weights_ok = False
                self.weights_err = f"Missing file: {WEIGHTS_FILE}"
                self.weights_cells_raw = {}
                self.weights_cells_parsed = {}
                return False

            wb, _ = _open_workbook(WEIGHTS_FILE)
            ws = wb[wb.sheetnames[0]]

            raw = {}
            parsed = {}

            for k in ("H3", "H4", "H5", "H6"):
                cell = WEIGHTS_FILE_CELLS[k]
                v = ws[cell].value
                raw[k] = v
                parsed[k] = to_date(v)

            for k in ("USD", "EUR", "NOK"):
                cell = WEIGHTS_FILE_CELLS[k]
                v = ws[cell].value
                raw[k] = v
                parsed[k] = safe_float(v, None)

            wb.close()

            self.weights_cells_raw = dict(raw)
            self.weights_cells_parsed = dict(parsed)
            self.weights_ok = True
            self.weights_err = None
            self.weights_last_loaded_ts = datetime.now()
            return True
        except Exception as e:
            self.weights_ok = False
            self.weights_err = str(e)
            self.weights_cells_raw = {}
            self.weights_cells_parsed = {}
            return False

    def load_recon_direct(self) -> tuple[bool, str]:
        try:
            file_path, msg = self.resolve_latest_path()
            if not file_path:
                return False, msg

            try:
                st = file_path.stat()
                self._last_src = file_path
                self._last_mtime = st.st_mtime
                self._last_size = st.st_size
                self._last_hash = compute_file_hash(file_path)

                # Register file with watcher for automatic change detection
                self._file_watcher.watch_file(file_path)
            except Exception:
                pass

            # Strategy 1: Try reading directly from running Excel via xlwings
            #   - Always fresh (reads from Excel's memory, not disk)
            #   - Works even if file is locked
            #   - Only available on Windows with Excel open
            xlw_result = _read_cells_xlwings(file_path, REQUIRED_CELLS, EXCEL_CM_RATES_MAPPING)

            if xlw_result is not None:
                recon, cm_rates, sheet_name = xlw_result
                used_cache = False
                read_source = "xlwings"
                log.info(f"[ExcelEngine.load_recon_direct] Read via xlwings (live Excel, sheet: {sheet_name})")
            else:
                # Strategy 2: Read from disk via openpyxl (with retry + cache fallback)
                wb, used_cache = _open_workbook(file_path)

                # Choose the correct sheet:
                # Prefer the latest date-formatted sheet if present.
                import re
                date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}$")
                date_sheets = [name for name in wb.sheetnames if date_pattern.match(str(name))]
                if date_sheets:
                    sheet_name = sorted(date_sheets)[-1]
                else:
                    sheet_name = wb.sheetnames[-1]
                ws = wb[sheet_name]

                recon = {}
                for (r, c) in REQUIRED_CELLS:
                    recon[(r, c)] = ws.cell(row=r, column=c).value

                cm_rates = {}
                for key, cell_ref in EXCEL_CM_RATES_MAPPING.items():
                    val = ws[cell_ref].value
                    cm_rates[key] = safe_float(val, None)

                wb.close()
                read_source = "cache copy" if used_cache else "openpyxl"
                log.info(f"[ExcelEngine.load_recon_direct] Read via {read_source} (sheet: {sheet_name})")

            self.recon_data = recon
            self.excel_cm_rates = cm_rates
            self.used_cache_fallback = used_cache
            log.info(f"[ExcelEngine.load_recon_direct] [OK] Excel CM rates loaded: {self.excel_cm_rates}")
            if used_cache:
                log.warning("[ExcelEngine.load_recon_direct] Data read from CACHE COPY — file was locked")
            self.last_loaded_ts = datetime.now()

            # Reset file watcher baseline so the next real change is detected
            if self._last_src and self._file_watcher:
                try:
                    self._file_watcher.force_rehash(self._last_src)
                except Exception:
                    pass

            self.load_weights_file()

            source_info = f"{self.current_year_loaded} / {self.current_filename}"
            if used_cache:
                source_info += " [CACHE — file was locked, data may be stale]"
            return True, source_info
        except Exception as e:
            return False, str(e)

    def check_changed(self) -> bool:
        """
        Check if Excel file has changed since last load.

        Uses MD5 hash comparison for reliable detection.
        """
        if not self._last_src:
            return False
        try:
            # Quick check: mtime or size changed?
            s = self._last_src.stat()
            if s.st_mtime == self._last_mtime and s.st_size == self._last_size:
                return False

            # Verify with hash
            current_hash = compute_file_hash(self._last_src)
            if current_hash and current_hash != self._last_hash:
                log.info(f"[ExcelEngine] File changed: {self._last_src.name}")
                return True
            return False
        except Exception:
            return False

    def on_change(self, callback: callable):
        """
        Register a callback to be notified when Excel file changes.

        The callback receives (file_path: Path) as argument.
        """
        if callback not in self._change_callbacks:
            self._change_callbacks.append(callback)
            log.info(f"[ExcelEngine] Registered change callback: {callback.__name__}")

    def remove_change_callback(self, callback: callable):
        """Remove a previously registered change callback."""
        if callback in self._change_callbacks:
            self._change_callbacks.remove(callback)

    def _on_file_changed(self, file_path: Path):
        """Internal callback from FileWatcher when file changes."""
        log.info(f"[ExcelEngine] FILE CHANGED DETECTED: {file_path.name}")

        # Notify all registered callbacks
        for callback in self._change_callbacks:
            try:
                callback(file_path)
            except Exception as e:
                log.error(f"[ExcelEngine] Change callback error: {e}")

    def stop_watching(self):
        """Stop the file watcher (call on application exit)."""
        if self._file_watcher:
            self._file_watcher.stop()
            log.info("[ExcelEngine] File watcher stopped")

    def get_days_for_date(self, date_str: str) -> dict:
        """
        Get days to maturity for all tenors on a specific date.
        
        Args:
            date_str: Date in format "YYYY-MM-DD"
            
        Returns:
            dict: {"1w_Days": 7, "1m_Days": 30, ...} or empty dict if not found
        """
        log.info(f"[ExcelEngine.get_days_for_date] Looking for date: {date_str}")
        
        if self.day_data.empty or "date" not in self.day_data.columns:
            log.info(f"[ExcelEngine.get_days_for_date] [ERROR] Day data not loaded")
            return {}
        
        try:
            # Convert input date to pandas Timestamp
            target_date = pd.Timestamp(date_str).normalize()
            log.info(f"[ExcelEngine.get_days_for_date] Target date: {target_date}")
            
            # Find matching row
            matching_rows = self.day_data[self.day_data["date"] == target_date]
            
            if matching_rows.empty:
                log.info(f"[ExcelEngine.get_days_for_date] [WARN] No data found for {date_str}")
                return {}
            
            # Get first matching row
            row = matching_rows.iloc[0]
            
            # Extract days columns
            days_map = {}
            for tenor in ["1w", "1m", "2m", "3m", "6m"]:
                days_col = f"{tenor}_Days"
                if days_col in row:
                    days_map[days_col] = row[days_col]
                    # Also add without "_Days" suffix for compatibility
                    days_map[tenor] = row[days_col]
            
            log.info(f"[ExcelEngine.get_days_for_date] [OK] Found days: {days_map}")
            return days_map
            
        except Exception as e:
            log.info(f"[ExcelEngine.get_days_for_date] [ERROR] {e}")
            return {}

    def get_future_days_data(self, limit_rows=300):
        """Get future NIBOR days data with improved debugging."""
        log.info(f"[ExcelEngine.get_future_days_data] Called with limit={limit_rows}")
        log.info(f"[ExcelEngine.get_future_days_data] day_data.empty={self.day_data.empty}")
        log.info(f"[ExcelEngine.get_future_days_data] day_data.shape={self.day_data.shape if not self.day_data.empty else 'N/A'}")
        
        if self.day_data.empty or "date" not in self.day_data.columns:
            log.info(f"[ExcelEngine.get_future_days_data] [ERROR] Returning empty DataFrame")
            if self._day_data_err:
                log.info(f"[ExcelEngine.get_future_days_data] Error was: {self._day_data_err}")
            return pd.DataFrame()
        
        today = pd.Timestamp(datetime.now().date()).normalize()
        log.info(f"[ExcelEngine.get_future_days_data] Today={today}")
        
        future_df = self.day_data[self.day_data["date"] >= today].copy()
        log.info(f"[ExcelEngine.get_future_days_data] Future rows: {len(future_df)}")
        
        for c in ["date", "settlement"]:
            if c in future_df.columns:
                future_df[c] = pd.to_datetime(future_df[c], errors="coerce").dt.strftime("%Y-%m-%d")
        
        future_df = future_df.reset_index(drop=True)
        if len(future_df) > limit_rows:
            future_df = future_df.iloc[:limit_rows].copy()
        
        log.info(f"[ExcelEngine.get_future_days_data] [OK] Returning {len(future_df)} rows")
        return future_df

    def get_recon_value(self, cell_ref: str) -> object | None:
        try:
            row, col = coordinate_to_tuple(cell_ref)
            return self.recon_data.get((row, col), None)
        except Exception:
            return None

    def get_internal_basket_rates(self) -> dict[str, float | None] | None:
        """
        Get Internal Basket Rates from latest sheet in Nibor workbook.
        
        Returns:
            dict: Rates mapped by tenor key (e.g., 'EUR_1M': 1.94)
            None: If file not found or error
        """

        import re

        # Use dynamic file lookup
        nibor_file = get_nibor_file_path()
        log.info(f"[ExcelEngine] Loading Internal Basket Rates from {nibor_file}")

        if not nibor_file.exists():
            log.info(f"[ExcelEngine] ERROR: Nibor workbook not found at {nibor_file}")
            return None

        try:
            wb, _ = _open_workbook(nibor_file)

            # Find latest sheet (YYYY-MM-DD format)
            date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
            date_sheets = [s.title for s in wb.worksheets if date_pattern.match(s.title)]

            if not date_sheets:
                log.error("[ExcelEngine] No date-formatted sheets found in workbook")
                wb.close()
                return None

            latest_sheet_name = sorted(date_sheets)[-1]
            latest_sheet = wb[latest_sheet_name]

            log.info(f"[ExcelEngine] Using sheet: {latest_sheet_name}")

            # Extract rates from cells
            rates = {}
            for key, cell_addr in INTERNAL_BASKET_MAPPING.items():
                value = latest_sheet[cell_addr].value
                rates[key] = safe_float(value, None)
                log.info(f"[ExcelEngine]   {key} ({cell_addr}): {rates[key]}")

            wb.close()
            return rates

        except (PermissionError, OSError, KeyError) as e:
            log.error(f"[ExcelEngine] Error loading Internal Basket Rates: {e}")
            return None

    def get_previous_sheet_nibor_rates(self) -> dict | None:
        """
        Get NIBOR rates from the SECOND-TO-LAST sheet in the workbook.

        Used for CHG calculation - compares current rates with previous saved sheet.
        Reads cells AA30-AA33 for 1M, 2M, 3M, 6M NIBOR rates.

        Returns:
            dict: {"1m": {"nibor": 4.52}, "2m": {...}, ...} or None if failed
            Also returns the sheet name (date) as "_date" key
        """
        import re

        # Use dynamic file lookup
        nibor_file = get_nibor_file_path()
        log.info("[ExcelEngine] Loading previous sheet NIBOR rates for CHG calculation...")
        log.info(f"[ExcelEngine] Workbook path: {nibor_file}")

        # Mapping: tenor -> cell address for NIBOR rates
        nibor_cell_mapping = {
            "1m": "AA30",
            "2m": "AA31",
            "3m": "AA32",
            "6m": "AA33"
        }

        # Check if file exists
        if not nibor_file.exists():
            log.info(f"[ExcelEngine] ERROR: Nibor workbook not found at {nibor_file}")
            return None

        try:
            wb, _ = _open_workbook(nibor_file)

            # Find date-formatted sheets (YYYY-MM-DD format)
            date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}$')
            date_sheets = [s.title for s in wb.worksheets if date_pattern.match(s.title)]

            if len(date_sheets) < 2:
                log.info(f"[ExcelEngine] Not enough sheets for CHG calculation (found {len(date_sheets)})")
                wb.close()
                return None

            # Get second-to-last sheet (sorted by date)
            sorted_sheets = sorted(date_sheets)
            prev_sheet_name = sorted_sheets[-2]  # Second to last
            prev_sheet = wb[prev_sheet_name]

            log.info(f"[ExcelEngine] Using previous sheet for CHG: {prev_sheet_name}")

            # Extract NIBOR rates from cells
            rates = {"_date": prev_sheet_name}
            for tenor, cell_addr in nibor_cell_mapping.items():
                value = prev_sheet[cell_addr].value
                nibor_value = safe_float(value, None)
                rates[tenor] = {"nibor": nibor_value}
                log.info(f"[ExcelEngine]   {tenor} ({cell_addr}): {nibor_value}")

            wb.close()
            return rates

        except Exception as e:
            log.info(f"[ExcelEngine] ERROR loading previous sheet NIBOR rates: {e}")
            return None

    def get_latest_weights(self, weights_path: Path) -> dict | None:
        """
        Get latest weights from Wheights.xlsx file.
        
        Always use the row with the latest date in column A.
        Uses columns F (USD), G (EUR), H (NOK).
        
        Args:
            weights_path: Path to Wheights.xlsx
        
        Returns:
            dict: {"USD": 0.88, "EUR": 0.12, "NOK": 0.00, "date": datetime} or None if failed
        """
        log.info("Loading latest weights from file...")
        log.info(f"[ExcelEngine] Weights file: {weights_path}")
        
        if not weights_path.exists():
            log.info(f"[ExcelEngine] [ERROR] Weights file not found: {weights_path}")
            return None
        
        try:
            wb, _ = _open_workbook(weights_path)
            ws = wb.active

            log.info(f"[ExcelEngine] Sheet: {ws.title}")


            date_col = WEIGHTS_COLS["Date"]
            usd_col = WEIGHTS_COLS["USD"]
            eur_col = WEIGHTS_COLS["EUR"]
            nok_col = WEIGHTS_COLS["NOK"]

            dated_rows = []

            for row in range(2, ws.max_row + 1):
                date_val = _parse_date_cell(ws.cell(row=row, column=date_col).value)
                if date_val:
                    dated_rows.append((date_val, row))

            if not dated_rows:
                log.error("[ExcelEngine] No valid dates found in column A")
                wb.close()
                return None

            dated_rows.sort(reverse=True)
            latest_date, latest_row = dated_rows[0]

            log.info(f"[ExcelEngine] Latest date: {latest_date.date()} (row {latest_row})")

            usd_weight = ws.cell(row=latest_row, column=usd_col).value
            eur_weight = ws.cell(row=latest_row, column=eur_col).value
            nok_weight = ws.cell(row=latest_row, column=nok_col).value

            try:
                weights = _parse_weights_row(usd_weight, eur_weight, nok_weight)
                weights["date"] = latest_date

                log.info(f"[ExcelEngine] Weights: USD={weights['USD']:.2%}, EUR={weights['EUR']:.2%}, NOK={weights['NOK']:.2%}")
                wb.close()
                return weights

            except (ValueError, TypeError) as e:
                log.error(f"[ExcelEngine] Cannot convert weights to float: {e}")
                wb.close()
                return None

        except Exception as e:
            log.error(f"[ExcelEngine] Failed to load weights: {e}", exc_info=True)
            return None
    
    def get_all_weights_history(self, weights_path: Path) -> list[dict]:
        """
        Get ALL weights history from Wheights.xlsx file.
        
        Returns list of all weights sorted by date (newest first).
        Uses columns: A (Date), F (USD), G (EUR), H (NOK).
        
        Args:
            weights_path: Path to Wheights.xlsx
        
        Returns:
            list of dicts: [{"date": datetime, "USD": 0.88, "EUR": 0.12, "NOK": 0.00}, ...]
            Returns empty list if failed
        """
        log.info("Loading ALL weights history from file...")
        log.info(f"[ExcelEngine] Weights file: {weights_path}")
        
        if not weights_path.exists():
            log.info(f"[ExcelEngine] [ERROR] Weights file not found: {weights_path}")
            return []
        
        try:
            wb, _ = _open_workbook(weights_path)
            ws = wb.active

            log.info(f"[ExcelEngine] Sheet: {ws.title}")


            date_col = WEIGHTS_COLS["Date"]
            usd_col = WEIGHTS_COLS["USD"]
            eur_col = WEIGHTS_COLS["EUR"]
            nok_col = WEIGHTS_COLS["NOK"]

            weights_history = []

            for row in range(2, ws.max_row + 1):
                date_val = _parse_date_cell(ws.cell(row=row, column=date_col).value)
                if not date_val:
                    continue

                try:
                    w = _parse_weights_row(
                        ws.cell(row=row, column=usd_col).value,
                        ws.cell(row=row, column=eur_col).value,
                        ws.cell(row=row, column=nok_col).value,
                    )
                    w["date"] = date_val
                    w["row"] = row
                    weights_history.append(w)
                except (ValueError, TypeError) as e:
                    log.warning(f"[ExcelEngine] Could not parse row {row}: {e}")
                    continue

            weights_history.sort(key=lambda x: x["date"], reverse=True)

            log.info(f"[ExcelEngine] Loaded {len(weights_history)} weight entries")
            if weights_history:
                latest = weights_history[0]
                log.info(f"[ExcelEngine] Latest: {latest['date'].date()} - USD={latest['USD']:.2%}, EUR={latest['EUR']:.2%}, NOK={latest['NOK']:.2%}")

            wb.close()
            return weights_history

        except Exception as e:
            log.error(f"[ExcelEngine] Failed to load weights history: {e}", exc_info=True)
            return []

    def write_confirmation_stamp(self) -> tuple[bool, str]:
        """
        Write confirmation stamp to Excel cells AE30-AE33.
        Format: ✓ CONFIRMED YYYY-MM-DD HH:MM by username

        Uses xlwings to write to open Excel file on Windows.
        Falls back to openpyxl if xlwings not available or file not open.

        Returns:
            tuple: (success: bool, message: str)
        """
        import getpass

        # Use dynamic file lookup
        nibor_file = get_nibor_file_path()

        # Create stamp text
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        user = getpass.getuser()
        stamp = f"✓ CONFIRMED {timestamp} by {user}"

        log.info(f"[ExcelEngine] Writing confirmation stamp to AE30-AE33: {stamp}")
        log.info(f"[ExcelEngine] Target file: {nibor_file}")

        # Try xlwings first (works with open Excel files on Windows)
        try:
            import xlwings as xw

            # Find the open workbook by name
            filename = nibor_file.name
            wb = None

            # Try to find the workbook in open Excel instances
            for app in xw.apps:
                for book in app.books:
                    if book.name == filename:
                        wb = book
                        break
                if wb:
                    break

            if wb is None:
                # Try to open it
                wb = xw.Book(str(nibor_file))

            # Get the last sheet
            ws = wb.sheets[-1]

            # Write stamp to cells AE30-AE33
            for row in range(30, 34):
                cell = ws.range(f"AE{row}")
                cell.value = stamp
                cell.color = (144, 238, 144)  # Light green RGB
                cell.api.Font.Bold = True
                cell.api.Font.Size = 9

            # Save the workbook
            wb.save()

            log.info(f"[ExcelEngine] Confirmation stamp written via xlwings")
            return True, f"Stamp written: {stamp}"

        except ImportError:
            log.info("[ExcelEngine] xlwings not available, trying openpyxl...")
        except Exception as e:
            log.warning(f"[ExcelEngine] xlwings failed: {e}, trying openpyxl...")

        # Fallback to openpyxl (requires file to be closed)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            if not nibor_file.exists():
                return False, f"Excel file not found: {nibor_file}"

            wb = load_workbook(nibor_file, data_only=False)
            ws = wb.worksheets[-1]

            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            font = Font(bold=True, size=9)

            for row in range(30, 34):
                cell = ws.cell(row=row, column=31)  # AE = column 31
                cell.value = stamp
                cell.fill = green_fill
                cell.font = font
                cell.alignment = Alignment(horizontal="left")

            wb.save(nibor_file)
            wb.close()

            log.info(f"[ExcelEngine] Confirmation stamp written via openpyxl")
            return True, f"Stamp written: {stamp}"

        except PermissionError:
            msg = "Excel file is open. Install xlwings (pip install xlwings) to write to open files."
            log.error(f"[ExcelEngine] {msg}")
            return False, msg
        except Exception as e:
            msg = f"Error writing stamp: {e}"
            log.error(f"[ExcelEngine] {msg}")
            return False, msg


def _load_mock_defaults_from_excel() -> dict:
    """
    Load mock Bloomberg data from Implied_NOK_Defaults.xlsx.
    Returns dict mapping ticker -> price.

    Includes:
    - Spot rates (F033 Curncy)
    - Forward prices (F033 Curncy)
    - CM rates (SWET Curncy)
    - Days to maturity (TPSF Curncy) - for DAYS_TO_MTY field
    """
    defaults_file = DATA_DIR / "Implied_NOK_Defaults.xlsx"

    # Fallback values if file doesn't exist
    fallback = {
        # Spots
        "NOK F033 Curncy": 10.20, "NKEU F033 Curncy": 11.95,
        # USDNOK forwards
        "NK1W F033 Curncy": 10.20, "NK1M F033 Curncy": 10.20,
        "NK2M F033 Curncy": 10.21, "NK3M F033 Curncy": 10.21,
        "NK6M F033 Curncy": 10.22,
        # EURNOK forwards
        "NKEU1W F033 Curncy": 11.96, "NKEU1M F033 Curncy": 11.98,
        "NKEU2M F033 Curncy": 11.99, "NKEU3M F033 Curncy": 12.01,
        "NKEU6M F033 Curncy": 12.07,
        # EUR CM rates
        "EUCM1M SWET Curncy": 1.8, "EUCM2M SWET Curncy": 1.9,
        "EUCM3M SWET Curncy": 2.02, "EUCM6M SWET Curncy": 2.11,
        # USD CM rates
        "USCM1M SWET Curncy": 3.65, "USCM2M SWET Curncy": 3.7,
        "USCM3M SWET Curncy": 3.73, "USCM6M SWET Curncy": 3.73,
        # NOK CM rates (NIBOR)
        "NKCM1M SWET Curncy": 4.6, "NKCM2M SWET Curncy": 4.55,
        "NKCM3M SWET Curncy": 4.5, "NKCM6M SWET Curncy": 4.4,
        # DAYS_TO_MTY - EURNOK (TPSF Curncy)
        "EURNOK1W TPSF Curncy": 7, "EURNOK1M TPSF Curncy": 30,
        "EURNOK2M TPSF Curncy": 60, "EURNOK3M TPSF Curncy": 90,
        "EURNOK6M TPSF Curncy": 180,
        # DAYS_TO_MTY - USDNOK (TPSF Curncy)
        "NK1W TPSF Curncy": 7, "NK1M TPSF Curncy": 30,
        "NK2M TPSF Curncy": 60, "NK3M TPSF Curncy": 90,
        "NK6M TPSF Curncy": 180,
    }

    if not defaults_file.exists():
        return fallback

    try:
        df = pd.read_excel(defaults_file, header=1)
        df.columns = ["TENOR", "EURNOK_SPOT", "EURNOK_PIPS", "USDNOK_SPOT",
                      "USDNOK_PIPS", "DAYS", "EUR_CM", "USD_CM", "NOK_CM"]

        prices = {}
        tenor_map = {"1W": "1W", "1M": "1M", "2M": "2M", "3M": "3M", "6M": "6M"}

        for _, row in df.iterrows():
            tenor = str(row["TENOR"]).strip().upper()
            if tenor not in tenor_map:
                continue

            t = tenor_map[tenor]

            # Spots (same for all tenors, use first row values)
            eur_spot = float(row["EURNOK_SPOT"])
            usd_spot = float(row["USDNOK_SPOT"])
            prices["NKEU F033 Curncy"] = eur_spot
            prices["NOK F033 Curncy"] = usd_spot

            # Forward prices = spot + pips/10000
            eur_pips = float(row["EURNOK_PIPS"])
            usd_pips = float(row["USDNOK_PIPS"])
            prices[f"NKEU{t} F033 Curncy"] = eur_spot + (eur_pips / 10000.0)
            prices[f"NK{t} F033 Curncy"] = usd_spot + (usd_pips / 10000.0)

            # CM rates
            prices[f"EUCM{t} SWET Curncy"] = float(row["EUR_CM"])
            prices[f"USCM{t} SWET Curncy"] = float(row["USD_CM"])
            prices[f"NKCM{t} SWET Curncy"] = float(row["NOK_CM"])

            # Days to maturity (DAYS_TO_MTY field from Bloomberg)
            days_val = int(row["DAYS"])
            prices[f"EURNOK{t} TPSF Curncy"] = days_val
            prices[f"NK{t} TPSF Curncy"] = days_val

        # Add 1W if not in file (estimate)
        if "NKEU1W F033 Curncy" not in prices:
            prices["NKEU1W F033 Curncy"] = prices.get("NKEU F033 Curncy", 11.95) + 0.005
        if "NK1W F033 Curncy" not in prices:
            prices["NK1W F033 Curncy"] = prices.get("NOK F033 Curncy", 10.20) + 0.0001
        if "EURNOK1W TPSF Curncy" not in prices:
            prices["EURNOK1W TPSF Curncy"] = 7
        if "NK1W TPSF Curncy" not in prices:
            prices["NK1W TPSF Curncy"] = 7

        # Duplicate F033 tickers to F043 for Dev/Prod mode compatibility
        prices = _add_dual_ticker_suffixes(prices)
        return prices

    except Exception as e:
        log.warning(f"Could not load mock defaults from Excel: {e}")
        return _add_dual_ticker_suffixes(fallback)


def _add_dual_ticker_suffixes(data: dict) -> dict:
    """Duplicate F033 tickers to F043 and vice versa for Dev/Prod mode compatibility."""
    result = dict(data)
    for ticker, value in list(data.items()):
        if "F033" in ticker:
            result[ticker.replace("F033", "F043")] = value
        elif "F043" in ticker:
            result[ticker.replace("F043", "F033")] = value
    return result


class BloombergEngine:
    """
    Bloomberg ingestion engine with caching.
    Automatically falls back to mock data if:
    - blpapi is not installed (Linux/cloud environment)
    - USE_MOCK_DATA is True in config
    """

    def __init__(self, cache_ttl_sec: float = 3.0):
        self._lock = threading.Lock()
        self._session = None
        self._service = None
        self._is_ready = False
        self._last_error = None

        self._cache_ttl_sec = float(cache_ttl_sec)
        self._cache_data: dict = {}
        self._cache_ts: float | None = None
        self._cache_tickers: tuple[str, ...] | None = None
        self._last_meta: dict = {}

        # NEVER use mock data if Bloomberg is available - it's dangerous in production
        # Only fall back to mock if blpapi is literally not installed
        self._use_mock = (blpapi is None)

        if self._use_mock:
            log.warning("[BloombergEngine] BLPAPI not installed - using mock data (DEV ONLY)")
            self._mock_prices = _load_mock_defaults_from_excel()
        else:
            log.info("[BloombergEngine] Using REAL Bloomberg data")
            self._mock_prices = {}
            self._start_session_async()

    def _start_session_async(self):
        def _starter():
            try:
                session_options = blpapi.SessionOptions()
                session_options.setServerHost("localhost")
                session_options.setServerPort(8194)
                session = blpapi.Session(session_options)
                if not session.start():
                    self._last_error = "Session Start Failed"
                    return
                if not session.openService("//blp/refdata"):
                    session.stop()
                    self._last_error = "Service Open Failed"
                    return
                service = session.getService("//blp/refdata")
                self._session = session
                self._service = service
                self._is_ready = True
                self._last_error = None
            except Exception as e:
                self._last_error = str(e)
        threading.Thread(target=_starter, daemon=True).start()

    def last_meta(self) -> dict:
        return dict(self._last_meta or {})

    def _generate_mock_price(self, ticker: str) -> dict:
        """Generate mock price data for a ticker from Excel defaults."""
        # Use exact price from Excel file (no randomization for verification)
        price = self._mock_prices.get(ticker, 1.0)

        # Generate realistic time string
        now = datetime.now()
        time_str = now.strftime("%H:%M:%S")

        return {"price": float(price), "change": 0.0, "time": time_str}

    def _fetch_mock_snapshot(self, tickers: list[str], callback_func, error_callback):
        """Fetch mock market data snapshot (used when blpapi is unavailable or USE_MOCK_DATA=True)."""
        tickers = [t for t in tickers if isinstance(t, str) and t.strip()]
        tickers_key = tuple(sorted(set(tickers)))

        now = time.time()
        # Check cache
        if self._cache_ts is not None and self._cache_tickers == tickers_key:
            age = now - float(self._cache_ts)
            if age <= self._cache_ttl_sec:
                meta = {
                    "request_id": uuid.uuid4().hex[:10],
                    "requested_at": datetime.fromtimestamp(now),
                    "received_at": datetime.fromtimestamp(now),
                    "duration_ms": 0,
                    "from_cache": True,
                    "requested_count": len(tickers_key),
                    "responded_count": len(self._cache_data),
                    "missing": [],
                    "mock": True,
                }
                self._last_meta = dict(meta)
                callback_func(dict(self._cache_data), dict(meta))
                return

        def _worker():
            with self._lock:
                req_id = uuid.uuid4().hex[:10]
                t0 = time.time()
                requested_at = datetime.fromtimestamp(t0)

                # Generate mock data for all tickers
                res = {}
                for ticker in tickers_key:
                    res[ticker] = self._generate_mock_price(ticker)

                # Simulate small network delay
                time.sleep(0.05)

                t1 = time.time()
                received_at = datetime.fromtimestamp(t1)
                duration_ms = int(round((t1 - t0) * 1000))

                meta = {
                    "request_id": req_id,
                    "requested_at": requested_at,
                    "received_at": received_at,
                    "duration_ms": duration_ms,
                    "from_cache": False,
                    "requested_count": len(tickers_key),
                    "responded_count": len(res),
                    "missing": [],
                    "mock": True,
                }
                self._last_meta = dict(meta)

                self._cache_data = dict(res)
                self._cache_ts = time.time()
                self._cache_tickers = tickers_key

                callback_func(res, meta)

        threading.Thread(target=_worker, daemon=True).start()

    def _ensure_ready_sync(self) -> tuple[bool, str | None]:
        if not blpapi:
            return False, "BLPAPI not installed"
        if self._is_ready and self._session and self._service:
            return True, None
        try:
            session_options = blpapi.SessionOptions()
            session_options.setServerHost("localhost")
            session_options.setServerPort(8194)
            session = blpapi.Session(session_options)
            if not session.start():
                return False, "Session Start Failed"
            if not session.openService("//blp/refdata"):
                session.stop()
                return False, "Service Open Failed"
            self._session = session
            self._service = session.getService("//blp/refdata")
            self._is_ready = True
            self._last_error = None
            return True, None
        except Exception as e:
            return False, str(e)

    def fetch_snapshot(self, tickers: list[str], callback_func, error_callback, fields: list[str] | None = None):
        # Only use mock if blpapi is literally not installed
        if self._use_mock:
            log.warning("[BloombergEngine] Using MOCK data - blpapi not installed!")
            self._fetch_mock_snapshot(tickers, callback_func, error_callback)
            return

        log.info(f"[BloombergEngine] Fetching REAL Bloomberg data for {len(tickers)} tickers")

        # Split tickers into TPSF (days) and regular (price) tickers
        tpsf_tickers = [t for t in tickers if "TPSF Curncy" in t]
        regular_tickers = [t for t in tickers if "TPSF Curncy" not in t]
        
        tickers = [t for t in tickers if isinstance(t, str) and t.strip()]
        tickers_key = tuple(sorted(set(tickers)))

        now = time.time()
        if self._cache_ts is not None and self._cache_tickers == tickers_key:
            age = now - float(self._cache_ts)
            if age <= self._cache_ttl_sec:
                meta = {
                    "request_id": uuid.uuid4().hex[:10],
                    "requested_at": datetime.fromtimestamp(now),
                    "received_at": datetime.fromtimestamp(now),
                    "duration_ms": 0,
                    "from_cache": True,
                    "requested_count": len(tickers_key),
                    "responded_count": len(self._cache_data),
                    "missing": sorted(list(set(tickers_key) - set(self._cache_data.keys()))),
                }
                self._last_meta = dict(meta)
                callback_func(dict(self._cache_data), dict(meta))
                return

        def _worker():
            with self._lock:
                ok, err = self._ensure_ready_sync()
                if not ok:
                    self._last_error = err
                    error_callback(err or "Unknown Bloomberg error")
                    return

                req_id = uuid.uuid4().hex[:10]
                t0 = time.time()
                requested_at = datetime.fromtimestamp(t0)

                try:
                    req = self._service.createRequest("ReferenceDataRequest")
                    for t in tickers_key:
                        req.getElement("securities").appendValue(t)
                    
                    # Request BOTH price fields AND days field for all tickers
                    req.getElement("fields").appendValue("PX_LAST")
                    req.getElement("fields").appendValue("CHG_NET_1D")
                    req.getElement("fields").appendValue("LAST_UPDATE")
                    req.getElement("fields").appendValue("DAYS_TO_MTY")  # CRITICAL for TPSF tickers!

                    self._session.sendRequest(req)

                    res = {}
                    responded = set()

                    while True:
                        ev = self._session.nextEvent(500)
                        if ev.eventType() in (blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE):
                            for msg in ev:
                                if not msg.hasElement("securityData"):
                                    continue
                                arr = msg.getElement("securityData")
                                for i in range(arr.numValues()):
                                    sec = arr.getValueAsElement(i)
                                    t = sec.getElementAsString("security")
                                    if sec.hasElement("securityError"):
                                        continue
                                    if not sec.hasElement("fieldData"):
                                        continue
                                    flds = sec.getElement("fieldData")

                                    # For TPSF tickers, use DAYS_TO_MTY as "price"
                                    if "TPSF Curncy" in t:
                                        price = flds.getElementAsFloat("DAYS_TO_MTY") if flds.hasElement("DAYS_TO_MTY") else 0.0
                                        change = 0.0  # Days don't have changes
                                        time_str = ""
                                    else:
                                        # For regular tickers, use PX_LAST as "price"
                                        price = flds.getElementAsFloat("PX_LAST") if flds.hasElement("PX_LAST") else 0.0
                                        change = flds.getElementAsFloat("CHG_NET_1D") if flds.hasElement("CHG_NET_1D") else 0.0
                                        time_str = flds.getElementAsString("LAST_UPDATE") if flds.hasElement("LAST_UPDATE") else ""

                                    res[t] = {"price": float(price), "change": float(change), "time": str(time_str)}
                                    responded.add(t)

                            if ev.eventType() == blpapi.Event.RESPONSE:
                                break
                        elif ev.eventType() == blpapi.Event.TIMEOUT:
                            continue

                    t1 = time.time()
                    received_at = datetime.fromtimestamp(t1)
                    duration_ms = int(round((t1 - t0) * 1000))

                    missing = sorted(list(set(tickers_key) - responded))

                    meta = {
                        "request_id": req_id,
                        "requested_at": requested_at,
                        "received_at": received_at,
                        "duration_ms": duration_ms,
                        "from_cache": False,
                        "requested_count": len(tickers_key),
                        "responded_count": len(responded),
                        "missing": missing,
                    }
                    self._last_meta = dict(meta)

                    self._cache_data = dict(res)
                    self._cache_ts = time.time()
                    self._cache_tickers = tickers_key

                    callback_func(res, meta)
                except Exception as e:
                    self._last_error = str(e)
                    error_callback(str(e))

        threading.Thread(target=_worker, daemon=True).start()

    def fetch_fixing_history(self, num_dates: int = 3) -> dict:
        """
        Fetch historical NIBOR fixings using Bloomberg BDH.

        Args:
            num_dates: Number of historical dates to fetch

        Returns:
            dict: {date_str: {tenor: rate, ...}, ...}
        """
        log.info(f"[BloombergEngine] fetch_fixing_history called for {num_dates} dates")
        log.info(f"[BloombergEngine] _use_mock={self._use_mock}, _is_ready={self._is_ready}, blpapi={'available' if blpapi else 'None'}")

        # If in mock mode, return empty to trigger Excel fallback
        if self._use_mock:
            log.info("[BloombergEngine] Mock mode - skipping BDH, will use Excel fallback")
            return {}

        if not blpapi:
            log.warning("[BloombergEngine] blpapi not available")
            return {}

        # Try to ensure session is ready
        try:
            with self._lock:
                ok, err = self._ensure_ready_sync()
                if not ok:
                    log.error(f"[BloombergEngine] BDH session not ready: {err}")
                    return {}
                log.info("[BloombergEngine] Session ready for BDH request")

                # Calculate date range (14 calendar days to ensure we get at least 3 business days)
                end_date = datetime.now()
                start_date = end_date - timedelta(days=14)

                log.info(f"[BloombergEngine] Fetching BDH for NIBOR fixings: {start_date.date()} to {end_date.date()}")

                # Create BDH request
                req = self._service.createRequest("HistoricalDataRequest")

                # Add all NIBOR fixing tickers
                for tenor, ticker in NIBOR_FIXING_TICKERS.items():
                    req.getElement("securities").appendValue(ticker)

                req.getElement("fields").appendValue("PX_LAST")
                req.set("startDate", start_date.strftime("%Y%m%d"))
                req.set("endDate", end_date.strftime("%Y%m%d"))
                req.set("periodicitySelection", "DAILY")

                self._session.sendRequest(req)

                # Collect results by date
                results_by_date = {}
                tenor_by_ticker = {v: k for k, v in NIBOR_FIXING_TICKERS.items()}

                while True:
                    ev = self._session.nextEvent(5000)
                    if ev.eventType() in (blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE):
                        for msg in ev:
                            if not msg.hasElement("securityData"):
                                continue

                            sec_data = msg.getElement("securityData")
                            ticker = sec_data.getElementAsString("security")
                            tenor = tenor_by_ticker.get(ticker)

                            if not tenor:
                                continue

                            if not sec_data.hasElement("fieldData"):
                                continue

                            field_data = sec_data.getElement("fieldData")

                            for i in range(field_data.numValues()):
                                row = field_data.getValueAsElement(i)
                                if row.hasElement("date") and row.hasElement("PX_LAST"):
                                    date_val = row.getElementAsDatetime("date")
                                    # blpapi.Datetime has year/month/day as properties, not methods
                                    date_str = f"{date_val.year}-{date_val.month:02d}-{date_val.day:02d}"
                                    price = row.getElementAsFloat("PX_LAST")

                                    if date_str not in results_by_date:
                                        results_by_date[date_str] = {}
                                    results_by_date[date_str][tenor] = price

                        if ev.eventType() == blpapi.Event.RESPONSE:
                            break
                    elif ev.eventType() == blpapi.Event.TIMEOUT:
                        log.warning("[BloombergEngine] BDH request timeout")
                        break

                # Filter to only include dates with all 5 tenors
                complete_dates = {}
                required_tenors = set(NIBOR_FIXING_TICKERS.keys())

                for date_str, rates in sorted(results_by_date.items(), reverse=True):
                    if set(rates.keys()) == required_tenors:
                        complete_dates[date_str] = rates
                        log.info(f"[BloombergEngine] BDH fixing for {date_str}: {rates}")
                        if len(complete_dates) >= num_dates:
                            break
                    else:
                        missing = required_tenors - set(rates.keys())
                        log.warning(f"[BloombergEngine] Skipping {date_str}: missing tenors {missing}")

                log.info(f"[BloombergEngine] BDH returned {len(complete_dates)} complete fixing dates")
                return complete_dates

        except Exception as e:
            log.error(f"[BloombergEngine] BDH request failed: {e}")
            return {}


class MockBloombergEngine:
    """
    Mock Bloomberg engine for testing without a real Bloomberg connection.
    Returns realistic random data for all tickers in MARKET_STRUCTURE.
    """

    def __init__(self, cache_ttl_sec: float = 3.0):
        self._lock = threading.Lock()
        self._is_ready = True
        self._last_error = None
        self._cache_ttl_sec = float(cache_ttl_sec)
        self._cache_data: dict = {}
        self._cache_ts: float | None = None
        self._cache_tickers: tuple[str, ...] | None = None
        self._last_meta: dict = {}

        # Realistic base prices for different ticker types
        self._base_prices = {
            # Spot rates
            "NOK F033 Curncy": 10.85,      # USDNOK around 10.85
            "NKEU F033 Curncy": 11.75,     # EURNOK around 11.75

            # USDNOK forwards (slightly higher than spot due to forward points)
            "NK1W F033 Curncy": 10.852,
            "NK1M F033 Curncy": 10.858,
            "NK2M F033 Curncy": 10.865,
            "NK3M F033 Curncy": 10.875,
            "NK6M F033 Curncy": 10.905,

            # EURNOK forwards
            "NKEU1W F033 Curncy": 11.752,
            "NKEU1M F033 Curncy": 11.758,
            "NKEU2M F033 Curncy": 11.765,
            "NKEU3M F033 Curncy": 11.775,
            "NKEU6M F033 Curncy": 11.805,

            # EUR CM curves (interest rates around 3-4%)
            "EUCM1M SWET Curncy": 3.85,
            "EUCM2M SWET Curncy": 3.78,
            "EUCM3M SWET Curncy": 3.72,
            "EUCM6M SWET Curncy": 3.55,

            # USD CM curves (interest rates around 4-5%)
            "USCM1M SWET Curncy": 4.85,
            "USCM2M SWET Curncy": 4.78,
            "USCM3M SWET Curncy": 4.72,
            "USCM6M SWET Curncy": 4.55,

            # NOK CM curves (NIBOR around 4-5%)
            "NKCM1M SWET Curncy": 4.65,
            "NKCM2M SWET Curncy": 4.58,
            "NKCM3M SWET Curncy": 4.52,
            "NKCM6M SWET Curncy": 4.35,
        }
        # Add F043 versions for Dev mode compatibility
        self._base_prices = _add_dual_ticker_suffixes(self._base_prices)

    def last_meta(self) -> dict:
        """Return metadata from the last fetch operation."""
        return dict(self._last_meta or {})

    def _generate_mock_price(self, ticker: str) -> dict:
        """Generate realistic mock price data for a ticker."""
        import random

        base = self._base_prices.get(ticker, 1.0)
        # Add small random variation (±0.5%)
        variation = base * random.uniform(-0.005, 0.005)
        price = round(base + variation, 4)

        # Random daily change (±0.3%)
        change = round(price * random.uniform(-0.003, 0.003), 4)

        # Generate realistic time string
        now = datetime.now()
        time_str = now.strftime("%H:%M:%S")

        return {"price": price, "change": change, "time": time_str}

    def fetch_snapshot(self, tickers: list[str], callback_func, error_callback, fields: list[str] | None = None):
        """
        Fetch mock market data snapshot for given tickers.

        Args:
            tickers: List of Bloomberg ticker strings
            callback_func: Function to call with (data_dict, meta_dict) on success
            error_callback: Function to call with error message on failure
            fields: Optional list of fields (ignored in mock, always returns price/change/time)
        """
        tickers = [t for t in tickers if isinstance(t, str) and t.strip()]
        tickers_key = tuple(sorted(set(tickers)))

        now = time.time()
        # Check cache
        if self._cache_ts is not None and self._cache_tickers == tickers_key:
            age = now - float(self._cache_ts)
            if age <= self._cache_ttl_sec:
                meta = {
                    "request_id": uuid.uuid4().hex[:10],
                    "requested_at": datetime.fromtimestamp(now),
                    "received_at": datetime.fromtimestamp(now),
                    "duration_ms": 0,
                    "from_cache": True,
                    "requested_count": len(tickers_key),
                    "responded_count": len(self._cache_data),
                    "missing": [],
                    "mock": True,
                }
                self._last_meta = dict(meta)
                callback_func(dict(self._cache_data), dict(meta))
                return

        def _worker():
            with self._lock:
                req_id = uuid.uuid4().hex[:10]
                t0 = time.time()
                requested_at = datetime.fromtimestamp(t0)

                # Generate mock data for all tickers
                res = {}
                for ticker in tickers_key:
                    res[ticker] = self._generate_mock_price(ticker)

                # Simulate small network delay
                time.sleep(0.05)

                t1 = time.time()
                received_at = datetime.fromtimestamp(t1)
                duration_ms = int(round((t1 - t0) * 1000))

                meta = {
                    "request_id": req_id,
                    "requested_at": requested_at,
                    "received_at": received_at,
                    "duration_ms": duration_ms,
                    "from_cache": False,
                    "requested_count": len(tickers_key),
                    "responded_count": len(res),
                    "missing": [],
                    "mock": True,
                }
                self._last_meta = dict(meta)

                self._cache_data = dict(res)
                self._cache_ts = time.time()
                self._cache_tickers = tickers_key

                callback_func(res, meta)

        threading.Thread(target=_worker, daemon=True).start()
